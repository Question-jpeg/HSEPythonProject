import time

import requests
from requests.models import PreparedRequest
from bs4 import BeautifulSoup
import openpyxl
from datetime import date
from configparser import ConfigParser
import math

from components.LoggerWindow import Logger

class Utils:
    def __init__(self, logger: Logger, cache_path: str):        
        self.logger = logger
        self.cache_path = cache_path

    def getParams(self, config):
        params = {}
        for option in config:
            if option.get('extra', False):
                for extra in option['extra']:
                    if extra['value']:
                        params[extra['urlParam']] = extra['value']
            if option.get('urlParam', False) and option.get('value', False):
                params[option['urlParam']] = option['value']
        
        return params

    def getResponse(self, url, params):
        self.logger.println('Произвожу запрос')

        req = PreparedRequest()
        req.prepare_url(url, params)
        while True:  
            time.sleep(3)                      
            response = requests.get(url=req.url.replace('%25', '%'))
            if response.ok:
                self.logger.println(f'Ответ получен [{response.status_code}]')
                break
        
        return response

    def retrieveCount(self, result):
        return int(result[result.find(' ')+1:result.rfind(' ')].replace(' ', ''))

    def getCountText(self, config, url):
        soup = BeautifulSoup(self.getResponse(url, self.getParams(config)).text, 'html.parser')
        summary = soup.select_one('div[data-name="SummaryHeader"]')
        return summary.text if summary != None else 'Найдено 0 объявлений'

    def getMiss(self, half, options, optimizations, url):
        self.logger.println('Проверяю цену: ' + optimizations[0]['value'])
        innerCount = self.retrieveCount(self.getCountText(options + optimizations, url))
        miss = half - innerCount
        self.logger.println(f'Погрешность: {miss}\n')

        return miss

    def getMissTuple(self, count, options, optimizations, url, isBuy):
        if count == 0:
            return [0, optimizations[0]['value']]
        if count < 200:
            self.logger.println(f'Количество объявлений < 200')
            self.logger.println(f'Произвожу расчёт по страницам\n')
            return self.pageBasedMissTupleGetter(options, optimizations, isBuy, url)
        return self.percentBasedMissTupleGetter(count, options, optimizations, url)

    def pageBasedMissTupleGetter(self, options, optimizations, isBuy, url):
        roundStep = int(optimizations[1]['value'])

        params = self.getParams(options)
        soup = BeautifulSoup(self.getResponse(url, params).text, 'html.parser')
        summary = soup.select_one('div[data-name="SummaryHeader"]')

        total_count = self.retrieveCount(summary.text)
        target_listing = round(total_count / 2)
        listings = soup.select('article[data-name="CardComponent"]')
        count_on_page_avg = len(listings)
        target_page = math.ceil(target_listing / count_on_page_avg)

        self.logger.println(f'Объявлений на странице: {count_on_page_avg}')
        self.logger.println(f'Искомая цена на странице № {target_page}\n' )

        if target_page != 1:
            params['p'] = target_page
            soup = BeautifulSoup(self.getResponse(url, params).text, 'html.parser')
            listings = soup.select('article[data-name="CardComponent"]')

        target_listing_index = target_listing - 1
        min_index_of_total_on_page = count_on_page_avg * (target_page-1)
        to_pick_index = target_listing_index - min_index_of_total_on_page

        if isBuy:
            price_text = listings[to_pick_index].select_one('p[data-mark="PriceInfo"]').text
            price_text = price_text[0:price_text.rfind(' ')+4] + ' Круто'
        else:
            price_text = listings[to_pick_index].select_one('span[data-mark="MainPrice"]').text
        price = self.retrieveCount(f"Цена: {price_text}")
        return [0, self.roundToStep(price, roundStep)]

    def percentBasedMissTupleGetter(self, count, options, optimizations, url):
        roundStep = int(optimizations[1]['value'])
        half = count // 2
        miss = self.getMiss(half, options, optimizations, url)
        history = {miss: optimizations[0]['value']}
        
        changes = 0
        toCheck = self.getToCheckPercentBased(count, miss, optimizations[0]['value'], roundStep, changes)
        isPositive = miss > 0

        while (toCheck != optimizations[0]['value']) and (toCheck not in list(history.values())):
            optimizations[0]['value'] = toCheck
            miss = self.getMiss(half, options, optimizations, url)
            history[miss] = toCheck

            if (miss > 0) != isPositive:
                isPositive = not isPositive
                changes += 1

            optimized = False
            for key in list(history.keys()):
                if (miss > 0 and key < 0) or (miss < 0 and key > 0):
                    ratio = abs(miss / key)
                    if ratio >= 0.9 and ratio <= 1.1:
                        toCheck = self.roundToStep((int(toCheck)+ int(history[key])) / 2, roundStep)
                        optimized = True
                        break
            
            if not optimized:
                toCheck = self.getToCheckPercentBased(count, miss, toCheck, roundStep, changes)

        min_key = min([abs(key) for key in list(history.keys())])
        return [min_key, history.get(min_key, False) or history.get(-min_key, False)]

    def getToCheckPercentBased(self, count, miss, currentPrice, roundStep, changesCount):
        if miss == 0:
            return currentPrice

        currentPrice = int(currentPrice)
        roundStep = int(roundStep)

        percent = miss / count
        percent /= 2**changesCount


        if percent > 0:
            percent = min(0.3, percent)
            return self.roundToStep(currentPrice*(1+percent), roundStep)
        else:
            percent = max(-0.3, percent)
            return self.roundToStep(currentPrice+currentPrice*percent, roundStep)

    def roundToStep(self, n, roundStep):
        i = int(n // roundStep)
        rem = n % roundStep
        if rem >= (roundStep // 2):
            i+= 1
        return str(roundStep*i)

    def saveToCache(self, config: ConfigParser, sectionName, propertyName, avg):
        config[sectionName][propertyName] = avg
        with open(self.cache_path, 'w', encoding="utf-8") as file:
            config.write(file)

class ExcelWorker:
    def __init__(self, fileName) -> None:
        self.workBook = openpyxl.load_workbook(fileName)
        self.fileName = fileName

    def createRow(self, options, sheetName):
        sheet = self.workBook[sheetName]
        max_row = sheet.max_row+1
        current_column = 2

        sheet.cell(row=max_row, column=current_column, value=date.today())
        for option in options:
            if not option.get('hide', False):
                current_column += 1
                value = option['value'] if option.get('label', None) == None else option['label']
                sheet.cell(row=max_row, column=current_column, value=value)

        self.current_column = current_column
        self.current_sheetName = sheetName

        self.save()

    def addToRow(self, count, avg):
        sheet = self.workBook[self.current_sheetName]
        max_row = sheet.max_row
        current_column = self.current_column
        sheet.cell(row=max_row, column=current_column+1, value=count)
        sheet.cell(row=max_row, column=current_column+2, value=avg)

        self.current_column = current_column+2

        self.save()

    def moveColumnsPointer(self, count):
        self.current_column = self.current_column + count

    def save(self):
        self.workBook.save(self.fileName)

    def close(self):
        self.workBook.close()

    


